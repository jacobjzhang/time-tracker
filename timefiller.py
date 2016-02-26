import re
import mechanize
from openpyxl import load_workbook
from xml.etree import ElementTree
import enterdata

# Importing the data values from the file
wb = load_workbook(filename = 'test-timetracker.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

hodes_branches = [["0"],["1"],["2"],["3"],["4"],["5"]]
# Loop through each row in the spreadsheet
for i in hodes_branches:
	for row in ws.iter_rows('A2:E6'):

	    temp_list = []
	    for cell in row:
	        # Insert values into array format for the function
	        temp_list.append(cell.value)

	    enterdata.enterEntry(i, temp_list[0],temp_list[1],temp_list[2],temp_list[3],temp_list[4])

